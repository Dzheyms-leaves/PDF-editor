"""Job pack assembly and cross-document batch operations."""

from __future__ import annotations

import io
import zipfile

import pytest

from src.assemble import batch, pack
from src.pdfcompat import pymupdf


def make_pdf(pages: int, label: str = "Doc", toc: bool = False) -> bytes:
    doc = pymupdf.open()
    for index in range(pages):
        page = doc.new_page()
        page.insert_text((72, 140), f"{label} page {index + 1}", fontsize=18)
    if toc and pages >= 2:
        doc.set_toc([[1, f"{label} intro", 1], [1, f"{label} detail", 2]])
    data = doc.tobytes()
    doc.close()
    return data


def upload(client, data: bytes, name: str) -> str:
    response = client.post("/api/documents",
                           files={"files": (name, data, "application/pdf")})
    assert response.status_code == 200, response.text
    return response.json()[0]["doc_id"]


# --------------------------------------------------------------- job packs

def test_pack_front_matter_offsets_every_reference():
    sources = [("Spec", make_pdf(3)), ("Board", make_pdf(2)), ("Cert", make_pdf(1))]
    data = pack.build_pack(sources, cover={"title": "Riverside", "revision": "C"})

    doc = pymupdf.open(stream=data)
    try:
        # Cover + one contents sheet + six body pages.
        assert doc.page_count == 8
        toc = doc.get_toc()
        assert toc[0] == [1, "Contents", 2]
        assert [row[2] for row in toc if row[0] == 1] == [2, 3, 6, 8]

        # Every contents row links to the page it names.
        links = [link for link in doc[1].get_links() if link["kind"] == pymupdf.LINK_GOTO]
        assert [link["page"] for link in links] == [2, 5, 7]
    finally:
        doc.close()


def test_pack_keeps_each_source_outline_rebased():
    sources = [("Spec", make_pdf(3, "Spec", toc=True)),
               ("Board", make_pdf(2, "Board", toc=True))]
    data = pack.build_pack(sources, cover=None, contents=False)
    doc = pymupdf.open(stream=data)
    try:
        toc = doc.get_toc()
        assert [row[1] for row in toc] == [
            "Spec", "Spec intro", "Spec detail", "Board", "Board intro", "Board detail"]
        # No front matter here, so the first source starts on page 1.
        assert toc[0][2] == 1
        assert toc[3][2] == 4
    finally:
        doc.close()


def test_pack_numbers_the_body_but_not_the_front_matter():
    data = pack.build_pack([("Spec", make_pdf(2))], cover={"title": "X"},
                           footer="AES-2481")
    doc = pymupdf.open(stream=data)
    try:
        assert "AES-2481" not in doc[0].get_text()      # cover stays clean
        assert "Page 1 of 2" in doc[2].get_text()
        assert "AES-2481" in doc[2].get_text()
        assert "Page 2 of 2" in doc[3].get_text()
    finally:
        doc.close()


def test_pack_outline_preview_matches_the_built_pack():
    rows = [("Spec", 3), ("Board", 2), ("Cert", 1)]
    preview = pack.pack_outline(rows)
    data = pack.build_pack([(title, make_pdf(pages)) for title, pages in rows],
                           cover={"title": "X"})
    doc = pymupdf.open(stream=data)
    try:
        starts = [row[2] for row in doc.get_toc() if row[1] != "Contents"]
    finally:
        doc.close()
    assert [row["start"] for row in preview] == starts


def test_pack_rejects_nothing_and_unreadable_input():
    with pytest.raises(ValueError, match="no documents"):
        pack.build_pack([])
    with pytest.raises(ValueError, match="not a readable PDF"):
        pack.build_pack([("Broken", b"%PDF-1.4 nope")])


def test_pack_survives_a_bad_cover_logo():
    data = pack.build_pack([("Spec", make_pdf(1))], cover={"title": "X"},
                           logo=b"not an image")
    assert data.startswith(b"%PDF")


# ---------------------------------------------------------------- batching

def test_stamp_can_skip_the_first_page():
    data = batch.stamp(make_pdf(3), footer="AES", skip_first=True)
    doc = pymupdf.open(stream=data)
    try:
        assert "AES" not in doc[0].get_text()
        assert "Page 1 of 2" in doc[1].get_text()
    finally:
        doc.close()


def test_stamp_rejects_an_unknown_position():
    with pytest.raises(ValueError, match="position"):
        batch.stamp(make_pdf(1), position="middle")


def test_watermark_needs_text():
    with pytest.raises(ValueError, match="needs some text"):
        batch.watermark(make_pdf(1), text="   ")


def test_rotate_must_be_square():
    with pytest.raises(ValueError, match="multiple of 90"):
        batch.rotate(make_pdf(1), degrees=45)


def test_split_every_n_and_by_ranges():
    data = make_pdf(5)
    names = [name for name, _b in batch.split(data, every=2, label="Job Report.pdf")]
    assert names == ["Job Report_p1-2.pdf", "Job Report_p3-4.pdf", "Job Report_p5.pdf"]

    parts = batch.split(data, ranges="1-2, 4", label="Job Report.pdf")
    assert [name for name, _b in parts] == ["Job Report_p1-2.pdf", "Job Report_p4.pdf"]
    first = pymupdf.open(stream=parts[0][1])
    try:
        assert first.page_count == 2
    finally:
        first.close()


def test_split_rejects_a_range_outside_the_document():
    with pytest.raises(ValueError, match="outside this document"):
        batch.split(make_pdf(3), ranges="9-12")
    with pytest.raises(ValueError, match="not a page range"):
        batch.split(make_pdf(3), ranges="a-b")


@pytest.mark.parametrize("pattern,expected", [
    ("{name}", "Job Report.pdf"),
    ("{nn}_{name}", "07_Job Report.pdf"),
    ("{project} {rev} {pages}p", "Riverside C 5p.pdf"),
    ("", "Job Report.pdf"),
])
def test_rename_patterns(pattern, expected):
    assert batch.rename(pattern, "Job Report.pdf", 7, 5, "Riverside", "C") == expected


def test_rename_strips_characters_a_filesystem_would_reject():
    assert batch.rename("{name}", 'a/b:c*d?.pdf', 1, 1) == "a-b-c-d-.pdf"


# -------------------------------------------------------------------- API

def test_pack_endpoint_matches_its_preview(client):
    ids = [upload(client, make_pdf(3), "Spec.pdf"),
           upload(client, make_pdf(2), "Board.pdf")]
    body = {"sources": [{"doc_id": i, "title": ""} for i in ids],
            "cover": {"enabled": True, "title": "Riverside"},
            "filename": "riverside-pack"}

    preview = client.post("/api/batch/pack/preview", json=body).json()
    assert preview["front_matter"] == 2
    assert preview["total_pages"] == 7

    response = client.post("/api/batch/pack", json=body)
    assert response.status_code == 200
    assert 'filename="riverside-pack.pdf"' in response.headers["content-disposition"]

    doc = pymupdf.open(stream=response.content)
    try:
        assert doc.page_count == preview["total_pages"]
    finally:
        doc.close()


def test_batch_run_in_place_is_undoable(client):
    doc_id = upload(client, make_pdf(2), "Spec.pdf")
    before = client.get(f"/api/documents/{doc_id}").json()
    assert before["can_undo"] is False

    response = client.post("/api/batch/run", json={
        "doc_ids": [doc_id], "operation": "watermark",
        "params": {"text": "DRAFT"}, "in_place": True})
    assert response.status_code == 200
    assert "1 document" in response.json()["message"]

    after = client.get(f"/api/documents/{doc_id}").json()
    assert after["can_undo"] is True
    assert client.post(f"/api/documents/{doc_id}/undo").status_code == 200


def test_batch_run_download_leaves_documents_alone(client):
    doc_id = upload(client, make_pdf(2), "Spec.pdf")
    response = client.post("/api/batch/run", json={
        "doc_ids": [doc_id], "operation": "scrub", "in_place": False})
    assert response.status_code == 200
    assert response.content.startswith(b"%PDF")     # one document, one PDF back
    assert client.get(f"/api/documents/{doc_id}").json()["can_undo"] is False


def test_batch_run_zips_several_documents(client):
    ids = [upload(client, make_pdf(2), "A.pdf"), upload(client, make_pdf(2), "B.pdf")]
    response = client.post("/api/batch/run", json={
        "doc_ids": ids, "operation": "optimise", "in_place": False})
    assert response.headers["content-type"] == "application/zip"
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        assert sorted(archive.namelist()) == ["A.pdf", "B.pdf"]


def test_batch_split_returns_every_part(client):
    ids = [upload(client, make_pdf(4), "A.pdf"), upload(client, make_pdf(3), "B.pdf")]
    response = client.post("/api/batch/split", json={"doc_ids": ids, "every": 2})
    with zipfile.ZipFile(io.BytesIO(response.content)) as archive:
        assert len(archive.namelist()) == 4          # 2 from A, 2 from B


def test_batch_rejects_an_empty_or_unknown_selection(client):
    assert client.post("/api/batch/run", json={
        "doc_ids": [], "operation": "scrub"}).status_code == 400
    assert client.post("/api/batch/run", json={
        "doc_ids": ["nope"], "operation": "scrub"}).status_code == 404


def test_batch_reports_a_bad_parameter(client):
    doc_id = upload(client, make_pdf(1), "A.pdf")
    response = client.post("/api/batch/run", json={
        "doc_ids": [doc_id], "operation": "watermark", "params": {"text": ""}})
    assert response.status_code == 400
    assert "text" in response.json()["detail"]


def test_rename_preview_follows_the_given_order(client):
    ids = [upload(client, make_pdf(1), "B.pdf"), upload(client, make_pdf(1), "A.pdf")]
    rows = client.post("/api/batch/rename/preview", json={
        "doc_ids": ids, "pattern": "{nn}_{name}"}).json()
    assert [row["to"] for row in rows] == ["01_B.pdf", "02_A.pdf"]


# --------------------------------------------------------------- workbooks

def _sheets(payload: bytes):
    import io as _io

    import openpyxl

    return openpyxl.load_workbook(_io.BytesIO(payload))


def test_pack_manifest_workbook_registers_every_document(client):
    ids = [upload(client, make_pdf(3), "Spec.pdf"),
           upload(client, make_pdf(2), "Board.pdf")]
    response = client.post("/api/batch/manifest", json={
        "sources": [{"doc_id": i, "title": ""} for i in ids],
        "pattern": "{nn}_{name}"})
    assert response.status_code == 200
    assert "pack-manifest.xlsx" in response.headers["content-disposition"]

    book = _sheets(response.content)
    assert book.sheetnames == ["Pack contents", "Rename plan"]

    rows = list(book["Pack contents"].iter_rows(values_only=True))
    assert rows[0] == ("#", "Title", "Pages", "Starts on page")
    assert rows[1] == (1, "Spec", 3, 3)          # after cover + contents
    assert rows[2] == (2, "Board", 2, 6)
    assert rows[-1][1:3] == ("Total pages", 7)

    plan = list(book["Rename plan"].iter_rows(min_row=2, values_only=True))
    assert plan == [("Spec.pdf", "01_Spec.pdf"), ("Board.pdf", "02_Board.pdf")]


def test_manifest_omits_the_rename_sheet_without_a_pattern(client):
    doc_id = upload(client, make_pdf(1), "Spec.pdf")
    response = client.post("/api/batch/manifest",
                           json={"sources": [{"doc_id": doc_id}]})
    assert _sheets(response.content).sheetnames == ["Pack contents"]
