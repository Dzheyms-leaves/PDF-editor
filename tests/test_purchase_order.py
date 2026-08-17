"""Purchase-order parsing, pinned against the two real supplier layouts.

Sample A is a Xero-style PO: right-aligned numeric columns, descriptions that
wrap onto continuation lines, item codes living inside the description cell,
and header fields printed *below* their labels.

Sample B is an Intelligent Control PO: a two-line table header ("DISC"/"%",
"COD"/"E"), a dedicated item-code column, and header fields to the *right* of
their labels.
"""

from __future__ import annotations

import pytest

from src.extract.purchase_order import (
    match_field,
    parse_purchase_order,
    split_code_from_description,
)
from src.extract.textgrid import PageGrid
from src.pdfcompat import pymupdf

MY_COMPANY = ["Automated Electrical Solutions", "AES"]


def _parse(path):
    doc = pymupdf.open(path)
    try:
        grids = [PageGrid.from_page(doc[i], i + 1) for i in range(doc.page_count)]
        return parse_purchase_order(grids, filename=path.name, my_company=MY_COMPANY)
    finally:
        doc.close()


# ----------------------------------------------------------------- sample A

@pytest.fixture(scope="module")
def po_a(samples):
    return _parse(samples / "po_sample_a.pdf")


def test_sample_a_header(po_a):
    header = po_a.header
    assert header.po_number == "PO-10297"
    assert header.order_date == "6 Aug 2026"
    assert header.reference == "J1706 Bosco"
    assert header.total == "830.48"
    assert header.tax == "75.51"
    assert header.currency == "AUD"


def test_sample_a_supplier_is_the_counterparty(po_a):
    """The party that is *not* us must be reported as the supplier."""
    assert po_a.header.supplier == "SMARTHOMEWORKS PTY LTD"


def test_sample_a_columns(po_a):
    assert po_a.columns == [
        "Description", "Quantity", "Unit Price", "Discount", "GST", "Amount AUD",
    ]


def test_sample_a_line_items(po_a):
    items = po_a.line_items
    assert len(items) == 3

    # A wrapped description must be folded into its own row, not split off.
    assert items[0].description == (
        "Dynalite Recessed 360 Degree Multifunction Sensor - RECESSED (DUS360-CR)"
    )
    assert items[0].quantity == "1.00"
    assert items[0].unit_price == "138.0045"
    assert items[0].total_price == "138.00"

    assert items[1].part_code == "DACM"
    assert items[1].quantity == "5.00"
    assert items[1].total_price == "413.90"

    # Leading code token pulled out of the description cell.
    assert items[2].part_code == "PD-PCN"
    assert items[2].quantity == "1.00"
    assert items[2].total_price == "278.58"


def test_sample_a_totals_row_is_not_an_item(po_a):
    for item in po_a.line_items:
        assert "TOTAL" not in (item.description or "").upper()


# ----------------------------------------------------------------- sample B

@pytest.fixture(scope="module")
def po_b(samples):
    return _parse(samples / "po_sample_b.pdf")


def test_sample_b_header(po_b):
    header = po_b.header
    assert header.po_number == "00111797"
    assert header.order_date == "26/07/2026"
    assert header.supplier == "Intelligent Control"
    assert header.total == "$3,452.67"
    assert header.tax == "$313.88"


def test_sample_b_wrapped_header_row(po_b):
    """'DISC'/'%' and 'COD'/'E' span two lines and must be recombined."""
    assert "DISC %" in po_b.columns
    assert "CODE" in po_b.columns
    assert "MY ITEM NO." in po_b.columns


def test_sample_b_line_item(po_b):
    assert len(po_b.line_items) == 1
    item = po_b.line_items[0]
    assert item.part_code == "DYN/DACMv2"
    assert item.description == "Dynat communication module"
    assert item.quantity == "45"
    assert item.unit_price == "$118.04"
    assert item.discount == "35%"
    assert item.total_price == "$3,452.67"
    # An unmapped column is kept rather than silently dropped.
    assert item.extra.get("code") == "GST"


def test_no_warnings_on_known_layouts(po_a, po_b):
    assert po_a.warnings == []
    assert po_b.warnings == []


# -------------------------------------------------------------- unit pieces

@pytest.mark.parametrize("header,expected", [
    ("Description", "description"),
    ("MY ITEM NO.", "part_code"),
    ("QTY", "quantity"),
    ("Unit Price", "unit_price"),   # must beat both "unit" and "price"
    ("UNIT", "unit"),
    ("EXTENDED", "total_price"),
    ("Amount AUD", "total_price"),
    ("GST", "tax"),
    ("DISC %", "discount"),
])
def test_column_keyword_mapping(header, expected):
    field, score = match_field(header)
    assert field == expected, f"{header!r} mapped to {field!r} (score {score})"


def test_bare_code_column_is_not_claimed_as_part_code():
    """'CODE' alone is ambiguous (tax code, GL code) and must not win."""
    field, _score = match_field("CODE")
    assert field != "part_code"


@pytest.mark.parametrize("description,code,rest", [
    ("PD-PCN DyNet PC Node Computer Adaptor", "PD-PCN", "DyNet PC Node Computer Adaptor"),
    ("DACM", "DACM", "DACM"),
    ("Sensor - RECESSED (DUS360-CR)", "DUS360-CR", "Sensor - RECESSED (DUS360-CR)"),
    ("LED downlight warm white", None, "LED downlight warm white"),
])
def test_split_code_from_description(description, code, rest):
    got_code, got_rest = split_code_from_description(description)
    assert got_code == code
    assert got_rest == rest


def test_export_round_trip(samples, po_a):
    from src import exporters

    csv_bytes = exporters.purchase_order_csv(po_a)
    assert b"PO-10297" in csv_bytes
    assert b"DACM" in csv_bytes

    xlsx = exporters.purchase_order_xlsx(po_a)
    assert xlsx[:2] == b"PK"  # a real zip container

    text = exporters.purchase_order_text(po_a)
    assert text.count("\n") >= len(po_a.line_items)


# --------------------------------------------------- multi-page continuation

def _multipage_pdf():
    """Page 1 has the table header, page 2 continues it, page 3 is prose."""
    from src.pdfcompat import pymupdf

    doc = pymupdf.open()
    page = doc.new_page()
    page.insert_text((40, 60), "Purchase Order Number", fontsize=10)
    page.insert_text((40, 76), "PO-55512", fontsize=10)
    page.insert_text((40, 200), "Description", fontsize=9)
    page.insert_text((300, 200), "Quantity", fontsize=9)
    page.insert_text((380, 200), "Unit Price", fontsize=9)
    page.insert_text((470, 200), "Amount", fontsize=9)
    for index, row in enumerate([
        ("DACM-A relay module", "2.00", "82.78", "165.56"),
        ("DUS360 sensor", "1.00", "138.00", "138.00"),
    ]):
        y = 230 + index * 22
        for x, value in zip((40, 300, 380, 470), row):
            page.insert_text((x, y), value, fontsize=9)

    second = doc.new_page()
    for index, row in enumerate([
        ("PD-PCN node adaptor", "3.00", "278.58", "835.74"),
        ("DTP170 panel", "4.00", "99.00", "396.00"),
    ]):
        y = 100 + index * 22
        for x, value in zip((40, 300, 380, 470), row):
            second.insert_text((x, y), value, fontsize=9)

    third = doc.new_page()
    third.insert_text((40, 100), "Terms and conditions of supply", fontsize=11)
    third.insert_text((40, 130), "Payment is due within 30 days of invoice date.", fontsize=9)
    third.insert_text((40, 150), "Goods remain our property until paid in full.", fontsize=9)

    data = doc.tobytes()
    doc.close()
    return data


@pytest.fixture(scope="module")
def po_multipage(tmp_path_factory):
    path = tmp_path_factory.mktemp("po") / "multi.pdf"
    path.write_bytes(_multipage_pdf())
    return _parse(path)


def test_table_continues_onto_a_page_with_no_header(po_multipage):
    assert len(po_multipage.line_items) == 4
    assert [i.part_code for i in po_multipage.line_items] == [
        "DACM-A", "DUS360", "PD-PCN", "DTP170",
    ]


def test_continuation_items_keep_their_page_number(po_multipage):
    assert [i.page for i in po_multipage.line_items] == [1, 1, 2, 2]


def test_line_numbers_run_across_pages(po_multipage):
    """Numbering must not restart at 1 on the continuation page."""
    assert [i.line_no for i in po_multipage.line_items] == [1, 2, 3, 4]


def test_prose_page_yields_no_line_items(po_multipage):
    for item in po_multipage.line_items:
        assert "Payment" not in (item.description or "")
        assert "property" not in (item.description or "")


@pytest.mark.parametrize("raw,expected", [
    ("$3,452.67", 3452.67),
    ("1.00", 1.0),
    ("35%", 0.35),
    ("(120.50)", -120.5),
    ("138.0045", 138.0045),
    ("", None),
    ("n/a", None),
    (None, None),
])
def test_to_number(raw, expected):
    from src.exporters import to_number
    assert to_number(raw) == expected


def test_xlsx_writes_real_numbers(po_a):
    """Excel must receive numbers it can total, not text."""
    import io

    from openpyxl import load_workbook

    from src import exporters

    book = load_workbook(io.BytesIO(exporters.purchase_order_xlsx(po_a)))
    sheet = book["Line items"]
    headings = [c.value for c in sheet[1]]
    qty_col = headings.index("Qty") + 1
    total_col = headings.index("Line total") + 1

    assert isinstance(sheet.cell(row=2, column=qty_col).value, (int, float))
    assert isinstance(sheet.cell(row=2, column=total_col).value, (int, float))
    assert sheet.cell(row=3, column=total_col).value == pytest.approx(413.90)
