"""Antumbra designer: part codes, geometry, spec sheet and job handoff."""

from __future__ import annotations

import json

import pytest

from src.designer import catalogue, icons, jobs, render
from src.models import PanelDesign


def design(**overrides) -> dict:
    base = {
        "design_id": "d1",
        "name": "Lobby keypad",
        "family": "B",
        "series": "P",
        "region": "A",
        "buttons": 6,
        "button_finish": "W",
        "rim_finish": "A",
        "backlight": "white",
        "quantity": 1,
        "engraving": [
            {"index": 0, "lines": ["WELCOME"], "icon": "welcome", "icon_side": "left"},
            {"index": 5, "lines": ["ALL", "OFF"], "icon": "all-off", "icon_side": "left"},
        ],
    }
    base.update(overrides)
    return base


# ------------------------------------------------------------- part codes

@pytest.mark.parametrize("family,series,region,buttons,button,rim,expected", [
    ("B", "P", "A", 6, "W", "A", "PA6BPA-WA"),
    ("B", "P", "A", 4, "W", "C", "PA4BPA-WC"),
    ("B", "P", "A", 6, "M", "A", "PA6BPA-MA"),
    ("B", "L", "E", 4, "W", "W", "PA4BLE-WW"),
    ("B", "P", "E", 2, "N", "C", "PA2BPE-NC"),
    ("T", "P", "A", 6, "G", "M", "PA6TPA-GM"),
    ("D", "P", "A", 0, "S", "W", "PADPA-SW"),      # display carries no count
])
def test_part_codes(family, series, region, buttons, button, rim, expected):
    assert catalogue.part_code(family, series, region, buttons, button, rim) == expected


def test_validate_rejects_impossible_combinations():
    with pytest.raises(ValueError, match="only made in"):
        catalogue.validate("D", "L", "A", 0, "W", "A")      # no Lite display
    with pytest.raises(ValueError, match="buttons"):
        catalogue.validate("B", "P", "A", 3, "W", "A")      # odd button count
    with pytest.raises(ValueError, match="button finish"):
        catalogue.validate("B", "P", "A", 6, "Z", "A")
    with pytest.raises(ValueError, match="rim finish"):
        catalogue.validate("B", "P", "A", 6, "W", "Z")


def test_chrome_is_a_rim_only_finish():
    assert "C" in {f.code for f in catalogue.RIM_FINISHES}
    assert "C" not in {f.code for f in catalogue.BUTTON_FINISHES}


# ---------------------------------------------------------------- geometry

def test_layout_fills_the_plate_without_overlap():
    layout = catalogue.layout("B", "P", "A", 6)
    assert (layout["width_mm"], layout["height_mm"]) == (75.0, 116.0)
    assert len(layout["buttons"]) == 6

    for button in layout["buttons"]:
        assert button["x"] >= layout["face"]["x"]
        assert button["y"] >= layout["face"]["y"]
        assert button["x"] + button["w"] <= layout["face"]["x"] + layout["face"]["w"] + 0.01
        assert button["y"] + button["h"] <= layout["face"]["y"] + layout["face"]["h"] + 0.01
        # The engraving area must sit clear of the indicator LED, on whichever
        # side of the panel that button's indicator lives.
        text = button["text"]
        if button["side"] == "right":
            assert text["x"] + text["w"] < button["led"]["cx"] - button["led"]["r"]
        else:
            assert text["x"] > button["led"]["cx"] + button["led"]["r"]
        assert text["w"] > 0

    # Two columns, three rows, no two buttons sharing a slot.
    positions = {(b["row"], b["column"]) for b in layout["buttons"]}
    assert len(positions) == 6


def test_the_right_hand_column_mirrors_the_left():
    """The indicator belongs at the outer edge of its own column.

    A dot drawn on the left of a right-hand button would be sitting in the
    middle of the panel, nowhere near the part it is meant to describe.
    """
    layout = catalogue.layout("B", "P", "A", 6)
    centre = layout["width_mm"] / 2

    for button in layout["buttons"]:
        led = button["led"]["cx"]
        if button["column"] == 0:
            assert button["side"] == "left"
            assert led < centre
            assert led - button["x"] < button["x"] + button["w"] - led
            assert button["text"]["align"] == "left"
        else:
            assert button["side"] == "right"
            assert led > centre
            assert button["x"] + button["w"] - led < led - button["x"]
            assert button["text"]["align"] == "right"

    # Mirrored, so a label has the same room to breathe in either column.
    widths = {round(b["text"]["w"], 2) for b in layout["buttons"]}
    assert len(widths) == 1


def test_a_two_button_panel_still_mirrors():
    layout = catalogue.layout("B", "P", "A", 2)
    left, right = layout["buttons"]
    assert left["side"] == "left" and right["side"] == "right"
    assert right["led"]["cx"] > left["led"]["cx"]


def test_fewer_buttons_means_taller_buttons():
    six = catalogue.layout("B", "P", "A", 6)["buttons"][0]["h"]
    two = catalogue.layout("B", "P", "A", 2)["buttons"][0]["h"]
    assert two > six


def test_european_plate_is_square_and_display_has_a_screen():
    layout = catalogue.layout("B", "P", "E", 4)
    assert layout["width_mm"] == layout["height_mm"] == 86.0

    display = catalogue.layout("D", "P", "A", 0)
    assert display["buttons"] == []
    assert display["screen"]["w"] > 0


def test_ink_flips_on_dark_finishes():
    assert catalogue.ink_for(catalogue.button_finish("W").hex) == "#2A2A26"
    assert catalogue.ink_for(catalogue.button_finish("N").hex) == "#F0EFEA"


# ------------------------------------------------------------------- icons

def test_every_icon_uses_known_primitives_inside_its_box():
    allowed = {"line", "poly", "fpoly", "circle", "disc"}
    assert icons.ICONS, "the icon library should not be empty"
    for icon in icons.catalogue():
        assert icon["group"] in icons.GROUPS
        for shape in icon["shapes"]:
            assert shape[0] in allowed
            if shape[0] == "line":
                numbers = shape[1:5]
            elif shape[0] in ("poly", "fpoly"):
                numbers = shape[1]
            else:
                numbers = [shape[1], shape[2]]
            assert all(-15 <= value <= 115 for value in numbers), icon["id"]


# -------------------------------------------------------------- spec sheet

def test_spec_sheet_reads_back_its_own_content():
    from src.pdfcompat import pymupdf

    data = render.spec_sheet([design()], "Riverside", "Stage 2", "Acme")
    assert data.startswith(b"%PDF")

    doc = pymupdf.open(stream=data)
    try:
        assert doc.page_count == 1
        text = doc[0].get_text()
    finally:
        doc.close()

    for expected in ("ANTUMBRA ENGRAVING SPECIFICATION", "PA6BPA-WA", "Lobby keypad",
                     "Riverside", "WELCOME", "ENGRAVING SCHEDULE", "75 x 116 mm"):
        assert expected in text, f"{expected!r} missing from the spec sheet"


def test_spec_sheet_is_one_page_per_panel():
    from src.pdfcompat import pymupdf

    data = render.spec_sheet([design(), design(design_id="d2", region="E", buttons=4)])
    doc = pymupdf.open(stream=data)
    try:
        assert doc.page_count == 2
    finally:
        doc.close()


def test_spec_sheet_rejects_an_empty_job():
    with pytest.raises(ValueError):
        render.spec_sheet([])


def test_every_catalogue_combination_renders():
    """A finish or family that cannot be drawn would be a broken order."""
    from src.pdfcompat import pymupdf

    designs = []
    for family in catalogue.FAMILIES:
        for series_code in family.series:
            for region in catalogue.REGIONS:
                counts = family.counts or (0,)
                designs.append(design(
                    design_id=f"{family.code}{series_code}{region.code}",
                    family=family.code, series=series_code, region=region.code,
                    buttons=counts[-1],
                ))
    data = render.spec_sheet(designs)
    doc = pymupdf.open(stream=data)
    try:
        assert doc.page_count == len(designs)
    finally:
        doc.close()


# ------------------------------------------------------------ panel queue

def test_designs_expand_into_one_queue_entry_per_physical_panel():
    model = PanelDesign(**design(quantity=3))
    entries = jobs.to_panel_entries([model])
    assert [e.name for e in entries] == [
        "Lobby keypad (1 of 3)", "Lobby keypad (2 of 3)", "Lobby keypad (3 of 3)"]

    labels = [label.text for row in entries[0].rows for label in row]
    assert labels == ["WELCOME", "ALL OFF"]
    assert all(entry.source_file == "PA6BPA-WA" for entry in entries)


def test_icon_only_buttons_still_carry_a_label():
    model = PanelDesign(**design(engraving=[
        {"index": 0, "lines": [], "icon": "fan"},
    ]))
    entries = jobs.to_panel_entries([model])
    assert entries[0].rows[0][0].text == "[Fan]"


def test_panels_without_engraving_are_not_queued():
    assert jobs.to_panel_entries([PanelDesign(**design(engraving=[]))]) == []


# -------------------------------------------------------------------- API

def test_catalogue_endpoint_describes_the_whole_range(client):
    body = client.get("/api/designer/catalogue").json()
    assert {f["code"] for f in body["families"]} == {"B", "T", "D"}
    assert {f["code"] for f in body["rim_finishes"]} == {"W", "M", "C", "A"}
    assert body["icons"] and body["icon_groups"]
    assert body["limits"]["max_lines"] == 2


def test_check_returns_code_geometry_and_warnings(client):
    body = client.post("/api/designer/check", json=design(engraving=[
        {"index": 0, "lines": ["A LABEL FAR TOO LONG FOR A BUTTON"], "icon": "nope"},
    ])).json()
    assert body["ok"] is True
    assert body["part_code"] == "PA6BPA-WA"
    assert body["product"] == "AntumbraButton"
    assert body["slots"] == 6
    assert len(body["layout"]["buttons"]) == 6
    assert any("longer than" in w for w in body["warnings"])
    assert any("unknown icon" in w for w in body["warnings"])


def test_check_reports_an_impossible_build_without_raising(client):
    body = client.post("/api/designer/check", json=design(family="D", series="L")).json()
    assert body["ok"] is False
    assert body["errors"]


@pytest.mark.parametrize("fmt,head", [
    ("pdf", b"%PDF"),
    ("csv", b"\xef\xbb\xbfPanel,"),
])
def test_export_formats(client, fmt, head):
    response = client.post("/api/designer/export",
                           json={"designs": [design()], "job_name": "riverside", "fmt": fmt})
    assert response.status_code == 200
    assert response.content.startswith(head)
    assert f"riverside.{fmt}" in response.headers["content-disposition"]


def test_export_json_round_trips_through_check(client):
    response = client.post("/api/designer/export",
                           json={"designs": [design()], "job_name": "riverside", "fmt": "json"})
    saved = json.loads(response.content)
    assert saved["job_name"] == "riverside"

    again = client.post("/api/designer/check", json=saved["designs"][0]).json()
    assert again["ok"] and again["part_code"] == "PA6BPA-WA"


def test_export_rejects_an_empty_job(client):
    response = client.post("/api/designer/export", json={"designs": [], "fmt": "pdf"})
    assert response.status_code == 400


def test_panels_endpoint_feeds_the_engraving_queue(client):
    response = client.post("/api/designer/panels", json={"designs": [design(quantity=2)]})
    assert response.status_code == 200
    assert len(response.json()) == 2

    empty = client.post("/api/designer/panels",
                        json={"designs": [design(engraving=[])]})
    assert empty.status_code == 400


# ------------------------------------------------------ bill of materials

def test_bom_groups_identical_configurations():
    from src.designer import bom
    from src.models import PanelDesign

    designs = [
        PanelDesign(**design(design_id="a", name="Suite A", buttons=4, region="E",
                             button_finish="N", rim_finish="C", quantity=12)),
        PanelDesign(**design(design_id="b", name="Suite B", buttons=4, region="E",
                             button_finish="N", rim_finish="C", quantity=8)),
        PanelDesign(**design(design_id="c", name="Lobby", quantity=4)),
    ]
    result = bom.build(designs, price_book={"PA4BPE-NC": 498.0, "PA6BPA-WA": 412.5,
                                            "ENGRAVING": 18.0})

    by_code = {line.part_code: line for line in result.lines}
    assert by_code["PA4BPE-NC"].quantity == 20          # the two suites merge
    assert by_code["PA4BPE-NC"].panels == ["Suite A", "Suite B"]
    assert by_code["PA6BPA-WA"].quantity == 4
    # Every panel here carries engraving, so the labour line counts all of them.
    assert by_code["ENGRAVING"].quantity == 24

    assert result.subtotal == round(20 * 498.0 + 4 * 412.5 + 24 * 18.0, 2)
    assert result.tax == round(result.subtotal * 0.10, 2)
    assert result.total == round(result.subtotal + result.tax, 2)


def test_bom_flags_a_part_with_no_rate_instead_of_pricing_it_free():
    from src.designer import bom
    from src.models import PanelDesign

    result = bom.build([PanelDesign(**design(quantity=3))], price_book={})
    line = result.lines[0]
    assert line.priced is False
    assert line.rate == 0.0
    assert result.unpriced == [line.part_code, "ENGRAVING"]


def test_bom_request_rates_beat_the_price_book():
    from src.designer import bom
    from src.models import PanelDesign

    result = bom.build([PanelDesign(**design())],
                       price_book={"PA6BPA-WA": 400.0},
                       overrides={"PA6BPA-WA": 350.0})
    assert result.lines[0].rate == 350.0


def test_bom_skips_the_engraving_line_for_a_blank_panel():
    from src.designer import bom
    from src.models import PanelDesign

    result = bom.build([PanelDesign(**design(engraving=[]))])
    assert [line.part_code for line in result.lines] == ["PA6BPA-WA"]


def test_bom_extras_join_the_order():
    from src.designer import bom
    from src.models import PanelDesign, QuoteLine

    result = bom.build([PanelDesign(**design())],
                       extras=[QuoteLine(description="Freight", quantity=1, rate=95),
                               QuoteLine(description="Ignored", quantity=0, rate=50)])
    assert [line.description for line in result.lines][-1] == "Freight"
    assert len(result.lines) == 3          # panel, engraving, freight


def test_quote_pdf_states_its_numbers(client):
    from src.pdfcompat import pymupdf

    response = client.post("/api/designer/quote", json={
        "designs": [design(quantity=2)], "job_name": "riverside",
        "client": "Acme Electrical", "reference": "Q-2481",
        "rates": {"PA6BPA-WA": 400.0, "ENGRAVING": 20.0}, "fmt": "pdf"})
    assert response.status_code == 200

    doc = pymupdf.open(stream=response.content)
    try:
        text = doc[0].get_text()
    finally:
        doc.close()
    for expected in ("QUOTATION", "PA6BPA-WA", "Acme Electrical", "Q-2481",
                     "840.00", "GST", "924.00"):
        assert expected in text, f"{expected!r} missing from the quote"


def test_quote_formats(client):
    for fmt, head in (("csv", b"\xef\xbb\xbfPart code,"), ("xlsx", b"PK")):
        response = client.post("/api/designer/quote", json={
            "designs": [design()], "job_name": "riverside", "fmt": fmt})
        assert response.status_code == 200
        assert response.content.startswith(head)


def test_bom_rejects_an_impossible_panel(client):
    response = client.post("/api/designer/bom",
                           json={"designs": [design(family="D", series="L")]})
    assert response.status_code == 400


# ------------------------------------------------------------- templates

def test_engraving_templates_round_trip(client):
    body = {"name": "Hotel suite", "slots": 4, "engraving": [
        {"index": 0, "lines": ["MASTER"], "icon": "bulb", "icon_side": "left"},
        {"index": 1, "lines": ["DO NOT", "DISTURB"], "icon": "dnd", "icon_side": "left"},
    ]}
    saved = client.post("/api/designer/templates", json=body).json()
    assert len(saved) == 1
    template_id = saved[0]["template_id"]
    assert template_id
    assert saved[0]["engraving"][1]["lines"] == ["DO NOT", "DISTURB"]

    assert client.get("/api/designer/templates").json()[0]["template_id"] == template_id

    # Saving with the same id replaces rather than duplicates.
    again = client.post("/api/designer/templates",
                        json={**body, "template_id": template_id, "name": "Suite v2"}).json()
    assert len(again) == 1
    assert again[0]["name"] == "Suite v2"

    assert client.delete(f"/api/designer/templates/{template_id}").json() == []
    assert client.delete(f"/api/designer/templates/{template_id}").status_code == 404


def test_template_needs_a_name(client):
    assert client.post("/api/designer/templates",
                       json={"name": "  ", "slots": 6}).status_code == 400


def test_engraving_workbook_has_a_sheet_for_each_end_of_the_job(client):
    import io as _io

    import openpyxl

    response = client.post("/api/designer/export", json={
        "designs": [design(quantity=4)], "job_name": "riverside", "fmt": "xlsx"})
    assert response.status_code == 200
    assert "riverside.xlsx" in response.headers["content-disposition"]

    book = openpyxl.load_workbook(_io.BytesIO(response.content))
    assert book.sheetnames == ["Engraving", "Panels"]

    engraving = list(book["Engraving"].iter_rows(values_only=True))
    assert engraving[0] == ("Panel", "Location", "Product code", "Position",
                            "Line 1", "Line 2", "Icon")
    # An empty location is an empty cell, which openpyxl reads back as None.
    assert engraving[1] == ("Lobby keypad", None, "PA6BPA-WA", 1, "WELCOME", None,
                            "Welcome")
    # Every position appears, engraved or not, so the sheet is a full schedule.
    assert len(engraving) == 1 + 6

    panels = list(book["Panels"].iter_rows(values_only=True))
    assert panels[1][0] == "Lobby keypad"
    assert panels[1][4] == "PA6BPA-WA"
    assert panels[1][7] == 4                     # quantity


# -------------------------------------------------------------- label style

def test_a_label_is_engraved_at_the_size_asked_for():
    sizes = render.fitted_sizes(design(text_size_mm=2.0, engraving=[
        {"index": 0, "lines": ["LOUNGE"]}]))
    assert sizes[0] == pytest.approx(2.0, abs=0.01)


def test_a_label_too_long_for_its_size_shrinks_rather_than_vanishes():
    """insert_textbox drops an overrunning line in full, so the size gives way."""
    asked = 4.0
    sizes = render.fitted_sizes(design(text_size_mm=asked, engraving=[
        {"index": 0, "lines": ["PLANT ROOM EXTRACT"]}]))
    assert 0 < sizes[0] < asked


def _label_fonts(data: bytes, label: str) -> set:
    from src.pdfcompat import pymupdf

    doc = pymupdf.open(stream=data)
    try:
        return {span["font"] for block in doc[0].get_text("dict")["blocks"]
                for line in block.get("lines", []) for span in line["spans"]
                if span["text"].strip() == label}
    finally:
        doc.close()


def test_the_chosen_font_reaches_the_page():
    panel = dict(engraving=[{"index": 0, "lines": ["LOUNGE"]}])
    serif = _label_fonts(render.spec_sheet([design(font="serif", **panel)]), "LOUNGE")
    sans = _label_fonts(render.spec_sheet([design(**panel)]), "LOUNGE")

    assert any("Times" in name for name in serif)
    assert not any("Times" in name for name in sans)


def test_an_unknown_font_is_refused_rather_than_quietly_swapped():
    with pytest.raises(ValueError, match="font"):
        catalogue.engraving_font("comic")
    with pytest.raises(ValueError, match="taller"):
        catalogue.text_size_mm(catalogue.MAX_TEXT_MM + 1)


def test_check_says_when_a_label_will_not_fit_at_the_size_asked_for(client):
    result = client.post("/api/designer/check", json=design(
        text_size_mm=4.0,
        engraving=[{"index": 0, "lines": ["PLANT ROOM EXTRACT"]}])).json()
    assert result["ok"]
    assert any("Position 1" in w and "4 mm" in w for w in result["warnings"])


def test_check_rejects_a_font_the_laser_has_never_heard_of(client):
    result = client.post("/api/designer/check", json=design(font="comic")).json()
    assert not result["ok"]


# --------------------------------------------------------------- order form

def _registration_box(page):
    """The plate outline the corner marks describe, in points."""
    from src.pdfcompat import pymupdf

    verticals, horizontals = [], []
    for drawing in page.get_drawings():
        for item in drawing["items"]:
            if item[0] != "l":
                continue
            start, end = item[1], item[2]
            if abs(abs(start.x - end.x) + abs(start.y - end.y)
                   - render.TICK_LEN) > 0.01:
                continue
            if abs(start.x - end.x) < 0.01:
                verticals.append(start.x)
            else:
                horizontals.append(start.y)
    return pymupdf.Rect(min(verticals), min(horizontals),
                        max(verticals), max(horizontals))


def test_order_form_marks_the_plate_at_full_size():
    """The marks are what the operator lines the part up against.

    Anything but 1:1 makes the sheet useless, and nothing on screen would say
    so, which is why the scale is pinned here.
    """
    from src.pdfcompat import pymupdf

    doc = pymupdf.open(stream=render.order_form([design()], "riverside"))
    try:
        page = doc[0]
        assert (round(page.rect.width), round(page.rect.height)) == (842, 595)
        box = _registration_box(page)
        assert box.width / render.MM == pytest.approx(75.0, abs=0.05)
        assert box.height / render.MM == pytest.approx(116.0, abs=0.05)
    finally:
        doc.close()


def test_order_form_engraves_each_label_inside_its_own_button():
    from src.pdfcompat import pymupdf

    panel = design(engraving=[
        {"index": 0, "lines": ["LOUNGE"]},
        {"index": 1, "lines": ["WINE ROOM"]},
        {"index": 5, "lines": ["ALL OFF"]},
    ])
    layout = catalogue.layout("B", "P", "A", 6)
    areas = {b["index"]: b["text"] for b in layout["buttons"]}

    doc = pymupdf.open(stream=render.order_form([panel]))
    try:
        page = doc[0]
        box = _registration_box(page)
        found = {}
        for block in page.get_text("dict")["blocks"]:
            for line in block.get("lines", []):
                for span in line["spans"]:
                    if span["bbox"][1] < box.y1:          # artwork, not the block
                        found[span["text"].strip()] = span["bbox"]
    finally:
        doc.close()

    assert set(found) == {"LOUNGE", "WINE ROOM", "ALL OFF"}
    for index, label in ((0, "LOUNGE"), (1, "WINE ROOM"), (5, "ALL OFF")):
        left, _top, right, _bottom = found[label]
        area = areas[index]
        assert (left - box.x0) / render.MM >= area["x"] - 0.2
        assert (right - box.x0) / render.MM <= area["x"] + area["w"] + 0.2


def test_order_form_labels_hug_the_indicator_in_either_column():
    """Mirrored on the page as on the part: left column reads from the left."""
    from src.pdfcompat import pymupdf

    panel = design(engraving=[{"index": 0, "lines": ["ON"]},
                              {"index": 1, "lines": ["ON"]}])
    doc = pymupdf.open(stream=render.order_form([panel]))
    try:
        page = doc[0]
        box = _registration_box(page)
        spans = sorted((span["bbox"] for block in page.get_text("dict")["blocks"]
                        for line in block.get("lines", []) for span in line["spans"]
                        if span["text"].strip() == "ON"
                        and span["bbox"][1] >= box.y0 and span["bbox"][3] <= box.y1),
                       key=lambda b: b[0])
    finally:
        doc.close()

    assert len(spans) == 2
    layout = catalogue.layout("B", "P", "A", 6)
    areas = {b["index"]: b["text"] for b in layout["buttons"]}
    # The left label starts at its area's left edge; the right one ends at its
    # area's right edge, which is where each column's indicator sits.
    assert (spans[0][0] - box.x0) / render.MM == pytest.approx(areas[0]["x"], abs=0.4)
    assert ((spans[1][2] - box.x0) / render.MM
            == pytest.approx(areas[1]["x"] + areas[1]["w"], abs=0.4))


def test_order_form_carries_the_panel_identity_and_its_engraving():
    from src.pdfcompat import pymupdf

    panel = design(name="lcp 1", order_12nc="913703057009")
    doc = pymupdf.open(stream=render.order_form([panel], "riverside"))
    try:
        text = doc[0].get_text()
    finally:
        doc.close()

    for expected in ("Product code: PA6BPA-WA", "12NC: 913703057009",
                     "Panel Name: lcp 1", "Orientation: Portrait",
                     "Style: American", "Type: Button", "Button Colour: White",
                     "ENGRAVING"):
        assert expected in text, f"{expected!r} missing from the order form"

    # The words appear twice over: cut into the artwork, and listed in the
    # block where the reference form prints a logo.
    assert text.count("WELCOME") == 2


def test_order_form_is_one_sheet_per_panel_and_needs_a_panel():
    from src.pdfcompat import pymupdf

    doc = pymupdf.open(stream=render.order_form(
        [design(), design(design_id="d2", region="E", buttons=4)]))
    try:
        assert doc.page_count == 2
        assert all(page.rect.width > page.rect.height for page in doc)
    finally:
        doc.close()

    with pytest.raises(ValueError):
        render.order_form([])


def test_order_form_download_is_named_for_the_job(client):
    response = client.post("/api/designer/export", json={
        "designs": [design()], "job_name": "riverside", "fmt": "order"})
    assert response.status_code == 200
    assert response.content.startswith(b"%PDF")
    assert "riverside-order-form.pdf" in response.headers["content-disposition"]
