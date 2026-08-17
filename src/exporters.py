"""CSV / TSV / XLSX / text exports for purchase orders and panel jobs."""

from __future__ import annotations

import csv
import io
from typing import Dict, List, Optional, Sequence

from .extract.panels import included_labels
from .models import PanelDesign, PanelEntry, PurchaseOrderResult

PO_FIELDS = [
    ("line_no", "Line"),
    ("part_code", "Part code"),
    ("description", "Description"),
    ("quantity", "Qty"),
    ("unit", "Unit"),
    ("unit_price", "Unit price"),
    ("discount", "Discount"),
    ("tax", "Tax"),
    ("total_price", "Line total"),
    ("due_date", "Due date"),
]


def _csv_bytes(rows: Sequence[Sequence[str]], delimiter: str = ",") -> bytes:
    buffer = io.StringIO(newline="")
    writer = csv.writer(buffer, delimiter=delimiter, lineterminator="\r\n")
    for row in rows:
        writer.writerow(["" if cell is None else str(cell) for cell in row])
    return buffer.getvalue().encode("utf-8-sig")


def _po_rows(result: PurchaseOrderResult, include_header_block: bool) -> List[List[str]]:
    extra_keys: List[str] = []
    for item in result.line_items:
        for key in item.extra:
            if key not in extra_keys:
                extra_keys.append(key)

    rows: List[List[str]] = []
    if include_header_block:
        header = result.header
        for label, value in (
            ("PO number", header.po_number),
            ("Order date", header.order_date),
            ("Required date", header.required_date),
            ("Supplier", header.supplier),
            ("Ship to", header.ship_to),
            ("Reference", header.reference),
            ("Currency", header.currency),
            ("Subtotal", header.subtotal),
            ("Tax", header.tax),
            ("Total", header.total),
        ):
            if value:
                rows.append([label, str(value)])
        rows.append([])

    rows.append([label for _key, label in PO_FIELDS] + extra_keys)
    for item in result.line_items:
        row = [getattr(item, key, "") or "" for key, _label in PO_FIELDS]
        row.extend(item.extra.get(key, "") for key in extra_keys)
        rows.append([str(cell) for cell in row])
    return rows


def purchase_order_csv(
    result: PurchaseOrderResult, include_header_block: bool = True, delimiter: str = ","
) -> bytes:
    return _csv_bytes(_po_rows(result, include_header_block), delimiter)


def purchase_order_xlsx(result: PurchaseOrderResult) -> bytes:
    """Two sheets: the order header, and the line items as a real table."""
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export") from exc

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Order"
    bold = Font(bold=True)

    header = result.header
    summary.append(["Field", "Value"])
    summary["A1"].font = bold
    summary["B1"].font = bold
    for label, value in (
        ("PO number", header.po_number),
        ("Order date", header.order_date),
        ("Required date", header.required_date),
        ("Supplier", header.supplier),
        ("Ship to", header.ship_to),
        ("Buyer", header.buyer),
        ("Reference", header.reference),
        ("Currency", header.currency),
        ("Subtotal", header.subtotal),
        ("Tax", header.tax),
        ("Total", header.total),
        ("Source file", result.filename),
        ("Read via", result.engine or result.source),
    ):
        if value:
            summary.append([label, str(value)])
    summary.column_dimensions["A"].width = 18
    summary.column_dimensions["B"].width = 52

    sheet = workbook.create_sheet("Line items")
    extra_keys: List[str] = []
    for item in result.line_items:
        for key in item.extra:
            if key not in extra_keys:
                extra_keys.append(key)

    headings = [label for _key, label in PO_FIELDS] + extra_keys
    sheet.append(headings)
    fill = PatternFill("solid", fgColor="FFEFD9BE")
    for column in range(1, len(headings) + 1):
        cell = sheet.cell(row=1, column=column)
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")

    for item in result.line_items:
        row = [getattr(item, key, "") or "" for key, _label in PO_FIELDS]
        row.extend(item.extra.get(key, "") for key in extra_keys)
        sheet.append(row)

    widths = [6, 18, 52, 8, 8, 12, 10, 8, 12, 12] + [14] * len(extra_keys)
    for index, width in enumerate(widths[: len(headings)], start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def purchase_order_text(result: PurchaseOrderResult) -> str:
    """Tab-separated text, sized for pasting straight into a spreadsheet."""
    lines: List[str] = []
    header = result.header
    if header.po_number:
        lines.append(f"PO {header.po_number}")
    if header.supplier:
        lines.append(header.supplier)
    if lines:
        lines.append("")
    lines.append("\t".join(label for _key, label in PO_FIELDS))
    for item in result.line_items:
        lines.append(
            "\t".join(str(getattr(item, key, "") or "") for key, _label in PO_FIELDS)
        )
    return "\n".join(lines)


# --------------------------------------------------------------------------
# Panel jobs
# --------------------------------------------------------------------------

def panels_csv(entries: Sequence[PanelEntry]) -> bytes:
    """One row per panel: name, then each button label in its own column.

    Matches the layout EZcad2's variable-text feature expects — point
    successive text objects at successive columns.
    """
    usable = [e for e in entries if not e.error and e.rows]
    max_labels = max((len(included_labels(e)) for e in usable), default=0)
    rows: List[List[str]] = [["Panel"] + [f"Label{i}" for i in range(1, max_labels + 1)]]
    for entry in usable:
        rows.append([entry.name, *included_labels(entry)])
    return _csv_bytes(rows)


def panels_text(entries: Sequence[PanelEntry]) -> str:
    blocks: List[str] = []
    for entry in entries:
        if entry.error or not entry.rows:
            continue
        lines = [f"Panel: {entry.name}", ""]
        for row in entry.rows:
            text = "     ".join(l.text for l in row if l.include)
            if text:
                lines.append(text)
        blocks.append("\n".join(lines))
    return "\n\n---\n\n".join(blocks)


def panels_xlsx(entries: Sequence[PanelEntry]) -> bytes:
    try:
        from openpyxl import Workbook  # noqa: PLC0415
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover
        raise RuntimeError("openpyxl is required for Excel export") from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Panels"
    usable = [e for e in entries if not e.error and e.rows]
    max_labels = max((len(included_labels(e)) for e in usable), default=0)

    headings = ["Panel"] + [f"Label{i}" for i in range(1, max_labels + 1)]
    sheet.append(headings)
    for column in range(1, len(headings) + 1):
        sheet.cell(row=1, column=column).font = Font(bold=True)
    for entry in usable:
        sheet.append([entry.name, *included_labels(entry)])

    sheet.column_dimensions["A"].width = 28
    for index in range(2, len(headings) + 1):
        sheet.column_dimensions[get_column_letter(index)].width = 20
    sheet.freeze_panes = "A2"

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# --------------------------------------------------------------------------
# Antumbra panel designs
# --------------------------------------------------------------------------

DESIGN_FIELDS = [
    "Panel",
    "Location",
    "Reference",
    "Product",
    "Product code",
    "Button finish",
    "Rim finish",
    "Quantity",
    "12NC",
    "Position",
    "Line 1",
    "Line 2",
    "Icon",
]


def designs_csv(designs: Sequence[PanelDesign]) -> bytes:
    """One row per engraved position, with each panel's order details repeated.

    Wide enough to drive the engraver and to be checked line by line against a
    purchase order, which is the form the workshop already reads.
    """
    from .designer import catalogue as _catalogue, icons as _icons

    rows: List[List[str]] = [list(DESIGN_FIELDS)]
    for design in designs:
        code = _catalogue.part_code(
            design.family, design.series, design.region, design.buttons,
            design.button_finish, design.rim_finish,
        )
        product = _catalogue.product_name(design.family, design.series)
        button = _catalogue.button_finish(design.button_finish).name
        rim = _catalogue.rim_finish(design.rim_finish).name
        slots = _catalogue.button_slots(design.family, design.buttons)
        by_index = {int(item.index): item for item in design.engraving}

        if slots <= 0:
            rows.append([design.name, design.location, design.reference, product,
                         code, button, rim, str(design.quantity),
                         design.order_12nc, "", "", "", ""])
            continue

        for index in range(slots):
            item = by_index.get(index)
            lines = [str(line).strip() for line in (item.lines if item else [])]
            lines += ["", ""]
            icon = _icons.get(getattr(item, "icon", "") or "") if item else None
            rows.append([
                design.name, design.location, design.reference, product, code,
                button, rim, str(design.quantity), design.order_12nc,
                str(index + 1), lines[0], lines[1], icon["name"] if icon else "",
            ])
    return _csv_bytes(rows)
